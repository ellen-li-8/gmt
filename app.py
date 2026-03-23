import os, re, io, time, logging
from flask import Flask, render_template, jsonify, send_file, Response
import requests
from bs4 import BeautifulSoup
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

app = Flask(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

TERM_YEARS = list(range(11, 26))  # 2011–2025 (SCOTUS uses 2-digit year)


# ─── SCOTUS scraping ──────────────────────────────────────────────────────────

def get_order_list_urls(term_year: int) -> list[dict]:
    """Fetch all Order List PDF URLs for a given 2-digit SCOTUS term year."""
    url = f"https://www.supremecourt.gov/orders/ordersofthecourt/{term_year:02d}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        log.warning(f"Failed to fetch order page for term {term_year}: {e}")
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    results = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)
        if "zor.pdf" in href.lower() or "order list" in text.lower():
            full_url = href if href.startswith("http") else "https://www.supremecourt.gov" + href
            date_match = re.search(r"(\d{2}/\d{2}/\d{2,4})", a.parent.get_text(" ") if a.parent else "")
            date_str = date_match.group(1) if date_match else ""
            results.append({"url": full_url, "date": date_str})
    return results


def parse_order_list_pdf(pdf_url: str, date_str: str, term_year: int) -> list[dict]:
    """Download a PDF and extract certiorari-granted cases."""
    try:
        r = requests.get(pdf_url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        log.warning(f"Failed to fetch PDF {pdf_url}: {e}")
        return []

    try:
        full_text = ""
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"
    except Exception as e:
        log.warning(f"Failed to parse PDF {pdf_url}: {e}")
        return []

    return extract_cases(full_text, date_str, term_year, pdf_url)


def extract_cases(text: str, date_str: str, term_year: int, source_url: str) -> list[dict]:
    """
    Parse raw PDF text and extract ALL cert petitions — both granted and denied.

    PDF sections look like:
      CERTIORARI GRANTED
        11-1234  SMITH V. JONES
      CERTIORARI DENIED
        11-9999  DOE V. ROE
        11-8888  FOO V. BAR
      HABEAS CORPUS DENIED
        ...
    We track which section we're in and tag each case with granted_cert = 1 or 0.
    """
    cases = []
    lines = text.split("\n")

    # Section states
    NONE = 0
    GRANTED = 1
    DENIED = 2
    current_section = NONE

    # Section header patterns
    granted_re  = re.compile(r"^\s*CERTIORARI\s+GRANTED\s*$", re.IGNORECASE)
    denied_re   = re.compile(r"^\s*CERTIORARI\s+DENIED\s*$", re.IGNORECASE)
    # Headers that end cert sections entirely
    other_section_re = re.compile(
        r"^\s*(HABEAS CORPUS|MANDAMUS|REHEARINGS|PROBABLE JURISDICTION NOTED|"
        r"AFFIRMED|REVERSED|DISMISSED|JUDGMENT|ORDERS IN PENDING|"
        r"CERTIORARI —|APPENDIX)\b",
        re.IGNORECASE
    )

    # Docket: e.g. "11-1234" or "11A234" at start of line
    docket_re = re.compile(r"^\s*(\d{1,2}[-A]\d{2,5})\s+(.*)")

    full_date = normalize_date(date_str, term_year)
    full_term = term_year_to_full(term_year, full_date)

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # Detect section changes
        if granted_re.match(stripped):
            current_section = GRANTED
            i += 1
            continue
        if denied_re.match(stripped):
            current_section = DENIED
            i += 1
            continue
        if other_section_re.match(stripped):
            current_section = NONE
            i += 1
            continue

        # Only parse dockets when inside a cert section
        if current_section in (GRANTED, DENIED):
            m = docket_re.match(line)
            if m:
                docket = m.group(1).strip()
                case_name_raw = m.group(2).strip()

                # Consume wrapped name lines: indented continuation, no docket start
                j = i + 1
                while j < len(lines):
                    next_stripped = lines[j].strip()
                    if not next_stripped:
                        break
                    if docket_re.match(lines[j]):
                        break
                    if granted_re.match(next_stripped) or denied_re.match(next_stripped):
                        break
                    if other_section_re.match(next_stripped):
                        break
                    # Stop at prose sentences (description of the ruling, not a case name)
                    if re.match(r"^(The |It |A |An |This |Petition|Motion|Justice)", next_stripped):
                        break
                    case_name_raw += " " + next_stripped
                    j += 1

                case_name = clean_case_name(case_name_raw)
                if case_name:
                    cases.append({
                        "docket": docket,
                        "case_name": case_name,
                        "date": full_date,
                        "term": full_term,
                        "granted_cert": 1 if current_section == GRANTED else 0,
                        "source_url": source_url,
                        "outcome": "",
                        "decision_direction": "",
                        "issue_area": "",
                        "oyez_url": "",
                        "circuit_split": 0,
                        "federalism": 0,
                        "precedent": 0,
                        "national_significance": 0,
                    })
                i = j
                continue

        i += 1

    return cases


def clean_case_name(raw: str) -> str:
    """Clean up extracted case name text."""
    # Remove trailing/leading noise
    name = re.sub(r"\s+", " ", raw).strip()
    # Strip docket references
    name = re.sub(r"\s*\d{1,2}[-A-Z]\d{3,5}\s*", " ", name).strip()
    # Strip parenthetical crud from consolidations
    name = re.sub(r"\s*[()]\s*$", "", name).strip()
    # Must have V. or v. to be a real case
    if not re.search(r"\bv\.?\s", name, re.IGNORECASE):
        return ""
    return name[:200]


def normalize_date(date_str: str, term_year: int) -> str:
    """Convert scraped date string to YYYY-MM-DD."""
    if not date_str:
        return ""
    # mm/dd/yy or mm/dd/yyyy
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", date_str)
    if m:
        mo, day, yr = m.groups()
        if len(yr) == 2:
            yr = "20" + yr
        return f"{yr}-{int(mo):02d}-{int(day):02d}"
    return date_str


def term_year_to_full(term_year: int, date_str: str) -> str:
    """Return human-readable term like 'October 2011'."""
    full_year = 2000 + term_year
    # SCOTUS term starts in October
    return f"October {full_year}"


# ─── Oyez enrichment ──────────────────────────────────────────────────────────

def enrich_with_oyez(cases: list[dict]) -> list[dict]:
    """
    Enrich cases with Oyez data.
    - Build a docket→oyez_summary lookup per term (fast, one call per term)
    - For GRANTED cases only, fetch the full case detail to get:
        outcome/winning_party, decision_direction, issue_area, decision date
    - For DENIED cases, just attach the oyez_url if found (no detail fetch needed)
    """
    # Step 1: build per-term docket lookup from Oyez list endpoint
    by_term: dict[str, list] = {}
    for c in cases:
        yr_match = re.search(r"(\d{4})", c.get("term", ""))
        if yr_match:
            by_term.setdefault(yr_match.group(1), []).append(c)

    # docket -> oyez summary object (has href, docket_number, name)
    oyez_summary: dict[str, dict] = {}

    for yr_str in by_term:
        log.info(f"Fetching Oyez term index for {yr_str}...")
        term_list = fetch_oyez_term(int(yr_str))
        for oc in term_list:
            docket = (oc.get("docket_number") or "").strip()
            if docket:
                oyez_summary[docket] = oc
                # Also index without leading term prefix, e.g. "1234" from "11-1234"
                alt = re.sub(r"^\d{2}-", "", docket)
                oyez_summary.setdefault(alt, oc)
        time.sleep(0.25)

    # Step 2: attach summary-level data to all cases
    for c in cases:
        oc = _find_oyez_match(c["docket"], oyez_summary)
        if oc:
            href = oc.get("href", "") or ""
            c["oyez_url"] = href.replace("api.oyez.org", "www.oyez.org")

    # Step 3: for GRANTED cases, fetch detail to get decisions + issue_area
    granted = [c for c in cases if c["granted_cert"] == 1]
    log.info(f"Fetching Oyez case details for {len(granted)} granted cases...")

    for idx, c in enumerate(granted):
        oc = _find_oyez_match(c["docket"], oyez_summary)
        if not oc:
            continue
        detail_url = oc.get("href", "")
        if not detail_url:
            continue
        detail = fetch_oyez_case_detail(detail_url)
        if not detail:
            continue

        # Decision direction + winning party
        decisions = detail.get("decisions") or []
        if decisions:
            d = decisions[0]
            c["decision_direction"] = d.get("decision_direction") or ""
            c["outcome"] = d.get("winning_party") or ""

        # Issue area
        c["issue_area"] = detail.get("issue_area") or ""

        # Decision date (more reliable than order list date)
        decision_date = detail.get("decision_date")
        if decision_date and not c["date"]:
            # Oyez stores as Unix timestamp
            try:
                from datetime import datetime, timezone
                dt = datetime.fromtimestamp(int(decision_date), tz=timezone.utc)
                c["date"] = dt.strftime("%Y-%m-%d")
            except Exception:
                pass

        # Argument date as fallback for the order date
        if not c["date"]:
            arg = detail.get("oral_argument_audio") or []
            if arg and isinstance(arg, list) and arg[0].get("href"):
                pass  # skip, too deep

        if idx % 20 == 0:
            log.info(f"  Oyez detail: {idx+1}/{len(granted)}")
        time.sleep(0.15)

    return cases


def _find_oyez_match(docket: str, lookup: dict) -> dict | None:
    """Try several docket key formats against the lookup table."""
    if not docket:
        return None
    # Direct match
    if docket in lookup:
        return lookup[docket]
    # Without leading zeros on the number part: "11-0123" -> "11-123"
    normalized = re.sub(r"-0+(\d)", r"-\1", docket)
    if normalized in lookup:
        return lookup[normalized]
    # Just the numeric suffix
    suffix = re.sub(r"^\d{2}-", "", docket)
    if suffix in lookup:
        return lookup[suffix]
    return None


def fetch_oyez_term(year: int) -> list[dict]:
    """Fetch all case summaries for a SCOTUS term year from Oyez API."""
    url = f"https://api.oyez.org/cases?filter=term:{year}&per_page=0"
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log.warning(f"Oyez term fetch failed for {year}: {e}")
        return []


def fetch_oyez_case_detail(href: str) -> dict | None:
    """Fetch a single case detail from Oyez API."""
    # href may be like https://api.oyez.org/cases/2011/10-1491
    try:
        r = requests.get(href, headers=HEADERS, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log.warning(f"Oyez detail fetch failed for {href}: {e}")
        return None


# ─── Excel export ─────────────────────────────────────────────────────────────

def build_excel(cases: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SCOTUS Cert Cases"

    headers = [
        "Case Name", "Docket #", "Date", "Term",
        "Granted Cert (0/1)", "Outcome / Winning Party",
        "Decision Direction", "Issue Area",
        "Circuit Split", "Federalism Conflict",
        "Precedent Matter", "National Significance",
        "Oyez URL", "Order List PDF",
    ]

    # Header row style
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="c8a96e", bold=True, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 36

    # Data rows
    row_fill_even = PatternFill(start_color="f5f3ee", end_color="f5f3ee", fill_type="solid")
    row_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx, c in enumerate(cases, 2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        row_data = [
            c.get("case_name", ""),
            c.get("docket", ""),
            c.get("date", ""),
            c.get("term", ""),
            c.get("granted_cert", 1),
            c.get("outcome", ""),
            c.get("decision_direction", ""),
            c.get("issue_area", ""),
            c.get("circuit_split", 0),
            c.get("federalism", 0),
            c.get("precedent", 0),
            c.get("national_significance", 0),
            c.get("oyez_url", ""),
            c.get("source_url", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx in (1, 6, 7, 8, 13, 14):
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Column widths
    col_widths = [45, 12, 14, 18, 16, 28, 22, 24, 14, 20, 18, 22, 40, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "SCOTUS Certiorari Dataset — Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri")
    granted_count = sum(1 for c in cases if c.get("granted_cert") == 1)
    denied_count  = sum(1 for c in cases if c.get("granted_cert") == 0)
    terms = sorted(set(c.get("term", "") for c in cases))
    ws2["A3"] = "Total cases:"
    ws2["B3"] = len(cases)
    ws2["A4"] = "Certiorari granted:"
    ws2["B4"] = granted_count
    ws2["A5"] = "Certiorari denied:"
    ws2["B5"] = denied_count
    ws2["A6"] = "Grant rate:"
    ws2["B6"] = f"{granted_count/len(cases)*100:.1f}%" if cases else "0%"
    ws2["A7"] = "Terms covered:"
    ws2["B7"] = ", ".join(terms)
    ws2["A8"] = "Generated:"
    from datetime import datetime
    ws2["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    for row in ws2.iter_rows(min_row=3, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── In-memory job store ──────────────────────────────────────────────────────
import uuid, threading

# job_id -> dict with keys: status, msg, pct, error, result_bytes, preview, total, granted, denied
_jobs: dict[str, dict] = {}


def _job_update(job_id: str, **kwargs) -> None:
    _jobs[job_id].update(kwargs)


# ─── Background scrape worker ─────────────────────────────────────────────────

def run_scrape_job(job_id: str) -> None:
    all_cases = []
    total_terms = len(TERM_YEARS)
    try:
        for idx, term_year in enumerate(TERM_YEARS):
            full_year = 2000 + term_year
            pct = int((idx / total_terms) * 50)
            _job_update(job_id, msg=f"Fetching order lists for {full_year} term ({idx+1}/{total_terms})...", pct=pct)

            order_urls = get_order_list_urls(term_year)
            if not order_urls:
                _job_update(job_id, msg=f"No order lists found for {full_year}, skipping.", pct=pct)
                continue

            _job_update(job_id, msg=f"Found {len(order_urls)} order lists for {full_year}. Parsing PDFs...", pct=pct)

            term_cases = []
            for pdf_info in order_urls:
                parsed = parse_order_list_pdf(pdf_info["url"], pdf_info["date"], term_year)
                term_cases.extend(parsed)
                time.sleep(0.05)

            seen: set[str] = set()
            unique = []
            for c in term_cases:
                key = f"{c['docket']}_{c['granted_cert']}"
                if key not in seen:
                    seen.add(key)
                    unique.append(c)

            all_cases.extend(unique)
            g = sum(1 for c in unique if c["granted_cert"] == 1)
            d = sum(1 for c in unique if c["granted_cert"] == 0)
            _job_update(job_id, msg=f"{full_year}: {len(unique)} cases ({g} granted, {d} denied).", pct=pct)

        _job_update(job_id, msg=f"Enriching {len(all_cases)} cases with Oyez data (this takes a few minutes)...", pct=52)
        all_cases = enrich_with_oyez(all_cases)

        _job_update(job_id, msg="Building Excel file...", pct=95)
        excel_bytes = build_excel(all_cases)

        granted_total = sum(1 for c in all_cases if c["granted_cert"] == 1)
        denied_total  = sum(1 for c in all_cases if c["granted_cert"] == 0)
        _job_update(
            job_id,
            status="done",
            msg=f"Done. {len(all_cases):,} cases — {granted_total} granted, {denied_total} denied.",
            pct=100,
            result_bytes=excel_bytes,
            preview=all_cases[:300],
            total=len(all_cases),
            granted=granted_total,
            denied=denied_total,
        )

    except Exception as e:
        log.exception("Scrape job failed")
        _job_update(job_id, status="error", msg=str(e), pct=0)


# ─── Flask routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/start", methods=["POST"])
def start():
    """Kick off a background scrape job. Returns job_id immediately."""
    job_id = str(uuid.uuid4())
    _jobs[job_id] = {"status": "running", "msg": "Starting...", "pct": 0}
    t = threading.Thread(target=run_scrape_job, args=(job_id,), daemon=True)
    t.start()
    return {"job_id": job_id}


@app.route("/api/status/<job_id>")
def status(job_id: str):
    """Poll for job progress. Returns JSON — safe short request, no timeout risk."""
    job = _jobs.get(job_id)
    if not job:
        return {"error": "Job not found"}, 404
    return {
        "status":  job.get("status", "running"),
        "msg":     job.get("msg", ""),
        "pct":     job.get("pct", 0),
        "total":   job.get("total", 0),
        "granted": job.get("granted", 0),
        "denied":  job.get("denied", 0),
        "preview": job.get("preview", []),
    }


@app.route("/api/download/<job_id>")
def download(job_id: str):
    """Serve the finished Excel file."""
    job = _jobs.get(job_id)
    if not job or job.get("status") != "done":
        return {"error": "Job not ready or not found"}, 404
    return send_file(
        io.BytesIO(job["result_bytes"]),
        download_name="scotus_cert_cases.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
